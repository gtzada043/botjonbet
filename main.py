import requests
import openpyxl
import os
import random
import string
import time
import tkinter as tk
from tkinter import messagebox
from playwright.sync_api import sync_playwright

# Função para criar planilha de contas
def criar_planilha_contas():
    if not os.path.exists("contasjonbet.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Contas"
        sheet.append(["Email", "Senha", "CPF", "Nome Completo", "CEP", "Endereco", "Cidade", "Estado"])  # Cabeçalhos das colunas
        workbook.save("contasjonbet.xlsx")
        print("Planilha 'contasjonbet.xlsx' criada. Adicione as informações necessárias na planilha.")
    else:
        print("Planilha 'contasjonbet.xlsx' já existe.")

# Função para criar planilha de proxies
def criar_planilha_proxies():
    if not os.path.exists("proxies.xlsx"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Proxies"
        sheet.append(["Proxy"])  # Cabeçalhos das colunas
        workbook.save("proxies.xlsx")
        print("Planilha 'proxies.xlsx' criada. Adicione as proxies na planilha.")
    else:
        print("Planilha 'proxies.xlsx' já existe.")

# Função para carregar proxies
def carregar_proxies():
    workbook = openpyxl.load_workbook("proxies.xlsx")
    sheet = workbook["Proxies"]
    proxies = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]]
    if not proxies:
        raise ValueError("Nenhuma proxy encontrada na planilha 'proxies.xlsx'.")
    return proxies

# Função para buscar dados do endereço via API ViaCEP
def buscar_dados_cep(cep):
    try:
        response = requests.get(f"https://viacep.com.br/ws/{cep}/json/")
        if response.status_code == 200:
            dados = response.json()
            return {
                "endereco": dados.get("logradouro", ""),
                "bairro": dados.get("bairro", ""),
                "cidade": dados.get("localidade", ""),
                "estado": dados.get("uf", ""),
                "pais": "Brasil"  # País sempre será Brasil
            }
        else:
            print(f"Erro ao buscar dados do CEP {cep}. Status code: {response.status_code}")
            return {}
    except Exception as e:
        print(f"Erro ao buscar dados do CEP: {e}")
        return {}

# Função para preencher os campos de endereço no formulário
def preencher_dados_endereco(page, cep, endereco, cidade, estado):
    try:
        # Esperar até que o campo de CEP esteja disponível
        page.wait_for_selector('input[data-testid="cep-input"]', timeout=60000)

        # Preenchendo os campos de endereço no formulário
        page.fill('input[data-testid="cep-input"]', cep)
        page.fill('input[data-testid="address-input"]', endereco)
        page.fill('input[data-testid="city-input"]', cidade)
        page.fill('input[data-testid="state-input"]', estado)

        # Preenchendo o país (sempre Brasil)
        page.select_option('select[data-testid="country-input"]', label="Brasil")

        print(f"Endereço preenchido com sucesso: {endereco}, {cidade}, {estado}")
    except Exception as e:
        print(f"Erro ao preencher os dados de endereço: {e}")

# Função para preencher o formulário de cadastro
def preencher_formulario(email, senha, cpf, nome_completo, url, cep, endereco, cidade, estado, proxy=None):
    with sync_playwright() as p:
        browser_args = {"headless": False}
        if proxy:
            proxy_parts = proxy.split(":")
            browser_args["proxy"] = {
                "server": f"{proxy_parts[0]}:{proxy_parts[1]}",
                "username": proxy_parts[2],
                "password": proxy_parts[3],
            }

        browser = p.chromium.launch(**browser_args)
        context = browser.new_context()
        page = context.new_page()

        try:
            acessar_pagina(page, url)
            print(f"Usando proxy: {proxy}")

            # Preenchendo os campos obrigatórios
            page.fill('input[data-testid="email-input"]', email)
            page.fill('input[data-testid="password-input"]', senha)
            page.fill('input[data-testid="national-id"]', cpf)

            # Preenchendo os campos de nome
            preencher_nome_sobrenome(page, nome_completo)

            # Preenchendo os dados de endereço
            preencher_dados_endereco(page, cep, endereco, cidade, estado)

            # Esperando 5 segundos para ver o que foi preenchido
            time.sleep(5)
        except Exception as e:
            print(f"Erro ao preencher o formulário para {email}: {e}")
        finally:
            page.close()
            context.close()
            browser.close()

# Função para acessar a página
def acessar_pagina(page, url):
    try:
        page.goto(url, timeout=50000)
        print(f"Página carregada com sucesso: {url}")
        time.sleep(3)
    except Exception as e:
        print(f"Erro ao acessar a página {url}: {e}")

# Função para preencher nome e sobrenome
def preencher_nome_sobrenome(page, nome_completo):
    try:
        primeiro_nome = nome_completo.split(" ")[0]
        ultimo_nome = nome_completo.split(" ")[-1]

        page.fill('input[data-testid="personal-info-first-name"]', primeiro_nome)
        page.fill('input[data-testid="personal-info-last-name"]', ultimo_nome)

        print(f"Nome preenchido: {primeiro_nome}, Sobrenome preenchido: {ultimo_nome}")
    except Exception as e:
        print("Erro ao preencher nome e sobrenome.", e)

# Função para iniciar o processo
def iniciar_processamento(instancias, url):
    proxies = carregar_proxies()
    workbook = openpyxl.load_workbook("contasjonbet.xlsx")
    sheet = workbook["Contas"]

    contas = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(row[:4]) and all(row[4:]):  # Verifica se os campos obrigatórios (exceto Senha) estão preenchidos
            email, senha, cpf, nome_completo, cep, endereco, cidade, estado = row
            senha = senha or gerar_senha()  # Gera a senha se estiver ausente
            contas.append((email, senha, formatar_cpf(cpf), nome_completo, cep, endereco, cidade, estado))
        else:
            print(f"Linha ignorada devido a dados faltantes: {row}")

    if not contas:
        raise ValueError("Nenhuma conta válida encontrada na planilha 'contasjonbet.xlsx'.")

    for i, (email, senha, cpf, nome_completo, cep, endereco, cidade, estado) in enumerate(contas):
        proxy = proxies[i % len(proxies)]
        print(f"Preenchendo formulário para: {email}")
        preencher_formulario(email, senha, cpf, nome_completo, url, cep, endereco, cidade, estado, proxy)

# Função para gerar senha
def gerar_senha(tamanho=12):
    return "Jongadas123."

# Função para formatar CPF
def formatar_cpf(cpf):
    cpf_str = str(cpf)
    return cpf_str.zfill(11)

# Função para mostrar a interface gráfica
def mostrar_interface():
    def on_submit():
        try:
            instancias = int(entry_instancias.get())
            url = entry_link.get()
            if instancias < 1 or not url:
                raise ValueError
            root.destroy()
            iniciar_processamento(instancias, url)
        except ValueError:
            messagebox.showerror("Erro", "Insira valores válidos para as instâncias e o link.")

    root = tk.Tk()
    root.title("BHP Automação")
    root.configure(bg='black')

    title_label = tk.Label(root, text="BHP Automação", fg="white", bg="black", font=("Helvetica", 16, "bold"))
    title_label.pack(pady=20)

    tk.Label(root, text="Quantidade de Instâncias Simultâneas:", fg="white", bg="black").pack(padx=10, pady=5)
    entry_instancias = tk.Entry(root)
    entry_instancias.pack(padx=10, pady=5)

    tk.Label(root, text="Link da Página:", fg="white", bg="black").pack(padx=10, pady=5)
    entry_link = tk.Entry(root)
    entry_link.pack(padx=10, pady=5)

    tk.Button(root, text="Iniciar", command=on_submit, bg="gray", fg="black").pack(padx=10, pady=10)

    root.mainloop()

# Função para criar planilhas
criar_planilha_contas()
criar_planilha_proxies()

# Inicia a interface
mostrar_interface()

